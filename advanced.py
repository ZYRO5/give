"""Extended utility functions for data processing and business logic."""

from datetime import datetime, timedelta
from typing import List, Dict, Tuple
import statistics


class DataProcessingUtilities:
    """Utilities for processing and analyzing data."""

    @staticmethod
    def calculate_statistics(data: List[float]) -> Dict[str, float]:
        """Calculate statistical metrics for numerical data."""
        if not data:
            return {}
        
        return {
            "mean": statistics.mean(data),
            "median": statistics.median(data),
            "mode": statistics.mode(data) if len(set(data)) < len(data) else None,
            "stdev": statistics.stdev(data) if len(data) > 1 else 0,
            "min": min(data),
            "max": max(data),
            "sum": sum(data),
            "count": len(data),
        }

    @staticmethod
    def calculate_percentile(data: List[float], percentile: int) -> float:
        """Calculate percentile value."""
        if not data or percentile < 0 or percentile > 100:
            return 0
        sorted_data = sorted(data)
        index = (len(sorted_data) - 1) * percentile / 100
        lower = int(index)
        upper = lower + 1
        if upper >= len(sorted_data):
            return sorted_data[lower]
        return sorted_data[lower] + (sorted_data[upper] - sorted_data[lower]) * (index - lower)

    @staticmethod
    def group_by_date(items: List[Dict], date_key: str, format: str = "%Y-%m-%d") -> Dict:
        """Group items by date."""
        grouped = {}
        for item in items:
            date = item.get(date_key)
            if date:
                if isinstance(date, str):
                    date = datetime.fromisoformat(date)
                date_str = date.strftime(format)
                if date_str not in grouped:
                    grouped[date_str] = []
                grouped[date_str].append(item)
        return grouped

    @staticmethod
    def flatten_list(nested_list: List[List]) -> List:
        """Flatten a nested list."""
        result = []
        for item in nested_list:
            if isinstance(item, list):
                result.extend(item)
            else:
                result.append(item)
        return result

    @staticmethod
    def chunk_list(lst: List, chunk_size: int) -> List[List]:
        """Split list into chunks."""
        return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

    @staticmethod
    def remove_duplicates(items: List[Dict], key: str) -> List[Dict]:
        """Remove duplicates from list based on key."""
        seen = set()
        result = []
        for item in items:
            value = item.get(key)
            if value not in seen:
                seen.add(value)
                result.append(item)
        return result

    @staticmethod
    def merge_dicts(dict_list: List[Dict]) -> Dict:
        """Merge multiple dictionaries."""
        result = {}
        for d in dict_list:
            result.update(d)
        return result

    @staticmethod
    def filter_by_date_range(items: List[Dict], date_key: str, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Filter items by date range."""
        result = []
        for item in items:
            date = item.get(date_key)
            if date:
                if isinstance(date, str):
                    date = datetime.fromisoformat(date)
                if start_date <= date <= end_date:
                    result.append(item)
        return result

    @staticmethod
    def sum_by_key(items: List[Dict], key: str) -> float:
        """Sum values by key."""
        return sum(item.get(key, 0) for item in items)

    @staticmethod
    def average_by_key(items: List[Dict], key: str) -> float:
        """Calculate average value by key."""
        values = [item.get(key, 0) for item in items if item.get(key)]
        return sum(values) / len(values) if values else 0

    @staticmethod
    def sort_by_key(items: List[Dict], key: str, reverse: bool = False) -> List[Dict]:
        """Sort items by key."""
        return sorted(items, key=lambda x: x.get(key, 0), reverse=reverse)

    @staticmethod
    def find_by_condition(items: List[Dict], condition: callable) -> List[Dict]:
        """Find items matching condition."""
        return [item for item in items if condition(item)]


class BusinessLogicUtilities:
    """Utilities for business logic calculations."""

    @staticmethod
    def calculate_donor_level(total_donated: float) -> str:
        """Determine donor level based on total amount donated."""
        if total_donated >= 100000:
            return "Platinum"
        elif total_donated >= 50000:
            return "Gold"
        elif total_donated >= 10000:
            return "Silver"
        elif total_donated >= 1000:
            return "Bronze"
        return "Friend"

    @staticmethod
    def calculate_campaign_progress_percentage(raised: float, target: float) -> float:
        """Calculate campaign progress percentage."""
        if target <= 0:
            return 0
        percentage = (raised / target) * 100
        return min(percentage, 100)

    @staticmethod
    def calculate_days_remaining(end_date: datetime) -> int:
        """Calculate days remaining until end date."""
        today = datetime.utcnow()
        delta = end_date - today
        return max(delta.days, 0)

    @staticmethod
    def is_campaign_active(start_date: datetime, end_date: datetime) -> bool:
        """Check if campaign is active."""
        now = datetime.utcnow()
        return start_date <= now <= end_date

    @staticmethod
    def calculate_donation_tax_deduction(amount: float, tax_rate: float = 0.20) -> float:
        """Calculate tax deduction for donation."""
        return amount * tax_rate

    @staticmethod
    def calculate_average_donation_value(total_donations: float, count: int) -> float:
        """Calculate average donation value."""
        return total_donations / count if count > 0 else 0

    @staticmethod
    def recommend_next_ask(total_donated: float, average_donation: float) -> float:
        """Recommend next donation ask based on history."""
        if average_donation <= 0:
            return 100
        return average_donation * 1.5  # Recommend 1.5x average

    @staticmethod
    def calculate_donor_lifetime_value(donations: List[float], retention_rate: float = 0.7) -> float:
        """Estimate donor lifetime value."""
        if not donations:
            return 0
        average = sum(donations) / len(donations)
        projected_donations = 10  # Assume 10 more donations
        return average * projected_donations * retention_rate

    @staticmethod
    def get_campaign_status_recommendation(raised: float, target: float, days_remaining: int) -> str:
        """Get recommendation for campaign status."""
        progress = (raised / target) * 100 if target > 0 else 0
        
        if progress >= 100:
            return "Complete"
        elif progress >= 75 and days_remaining > 7:
            return "On Track"
        elif progress >= 50 and days_remaining > 3:
            return "Needs Attention"
        else:
            return "Urgent"


class CacheKeyUtilities:
    """Utilities for cache key generation."""

    @staticmethod
    def generate_campaign_key(campaign_id: str) -> str:
        """Generate cache key for campaign."""
        return f"campaign:{campaign_id}"

    @staticmethod
    def generate_donor_key(donor_id: str) -> str:
        """Generate cache key for donor."""
        return f"donor:{donor_id}"

    @staticmethod
    def generate_donation_key(donation_id: str) -> str:
        """Generate cache key for donation."""
        return f"donation:{donation_id}"

    @staticmethod
    def generate_user_key(user_id: str) -> str:
        """Generate cache key for user."""
        return f"user:{user_id}"

    @staticmethod
    def generate_report_key(report_type: str, period: str) -> str:
        """Generate cache key for report."""
        return f"report:{report_type}:{period}"

    @staticmethod
    def generate_analytics_key(analytics_type: str, entity_id: str) -> str:
        """Generate cache key for analytics."""
        return f"analytics:{analytics_type}:{entity_id}"


class ExportUtilities:
    """Utilities for data export."""

    @staticmethod
    def export_to_csv_format(data: List[Dict], filename: str = None) -> str:
        """Export data to CSV format."""
        if not data:
            return ""
        
        import csv
        import io
        
        output = io.StringIO()
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue()

    @staticmethod
    def export_to_json_format(data: any) -> str:
        """Export data to JSON format."""
        import json
        from datetime import datetime
        
        def datetime_handler(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
        
        return json.dumps(data, default=datetime_handler, indent=2)

    @staticmethod
    def export_to_excel_format(data: List[Dict]) -> bytes:
        """Export data to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if data:
                headers = list(data[0].keys())
                ws.append(headers)
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row in data:
                    ws.append([row.get(h) for h in headers])
            
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:
            return b""
